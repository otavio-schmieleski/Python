# openpyxl 
from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent # nome do caminho raiz
WOOKBOOK_PATH = Path(__file__).parent / 'workbook.xlsx' # caminho e o nome do arquivo

wookbook = Workbook() # cria a planilha
worksheet: Worksheet= wookbook.active # esta é a planilha ativada

#ou assim 

# dando nome a planilha
# shett_name = 'minha planilha'
# criando a planilha
# Workbook.create_sheet(shett_name,0)
# selecinou a planilha 
# worksheet: worksheet = wookbook[shett_name]

# criar cabechalhos
worksheet.cell(1,1, 'Nome') # passa primeiro a linha, coluna e declaracao
worksheet.cell(1,2,'idade') # passa primeiro a linha, coluna e declaracao
worksheet.cell(1,3,'nota') # passa primeiro a linha, coluna e declaracao

students = [
    # nome  idade  nota
    ['joao', 14 ,  5.5],
    ['otavio', 18 , 10],
    ['luiz', 15,   12]
]

for student in students: # coloca tudo na planilha com append
    worksheet.append(student)

wookbook.save(WOOKBOOK_PATH) # para salvar na pasta com o nome do arquivo


"""LER O EXCEL """
# openpyxl - ler e alterar dados de uma planilha
# Com essa biblioteca será possível ler e escrever dados em células
# específicas, formatar células, inserir gráficos,
# criar fórmulas, adicionar imagens e outros elementos gráficos às suas
# planilhas. Ela é útil para automatizar tarefas envolvendo planilhas do
# Excel, como a criação de relatórios e análise de dados e/ou facilitando a
# manipulação de grandes quantidades de informações.
# Instalação necessária: pip install openpyxl
# Documentação: https://openpyxl.readthedocs.io/en/stable/
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

# Carregando um arquivo do excel
workbook: Workbook = load_workbook(WORKBOOK_PATH)

# Nome para a planilha
sheet_name = 'Minha planilha'

# Selecionou a planilha
worksheet: Worksheet = workbook[sheet_name]

row: tuple[Cell]
for row in worksheet.iter_rows(min_row=2):
    for cell in row:
        print(cell.value, end='\t')

        if cell.value == 'Maria':
            worksheet.cell(cell.row, 2, 23)
    print()

# worksheet['B3'].value = 14

workbook.save(WORKBOOK_PATH)